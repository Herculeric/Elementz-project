from requests_html import HTMLSession
from openpyxl import Workbook, load_workbook

#Variables
first_idx = 1
last_idx = 1000
column_to_write = 21 #Column "A" is 1, "B" is 2... "Z" is 26
file_name = r"C:\Users\ASUS\Desktop\Data collection\tracks.xlsx"
#Functions
def returnGenres(id):
    try:
        s = HTMLSession()
        response = s.get(f"https://www.chosic.com/music-genre-finder/?track={id}")
        response.html.render(sleep = 1,timeout=60)
        about = response.html.find('.pl-tags.tagcloud', first=True)
        s.close()
        return about.text
    except:
        return "None"

def isPopSong(genre):
    if "pop" in genre.lower():
        return 1
    else:
        return 0

def writeToCell(worksheet_reference,row_idx, genre_boolean,column_idx):
    worksheet_reference.cell(row=row_idx, column=column_idx).value = genre_boolean

#Execution
print("Opening Workbook, may take awhile...")
wb_read = Workbook()
wb_read = load_workbook(filename=file_name,read_only=True)
ws_read = wb_read["Sheet1"]
idList = []
resList = []
print("Getting IDs...")
rows = list(ws_read.rows)
print("Getting Genres...")

for cells in rows[first_idx:last_idx+1]:
    idList.append(cells[0].value)

for idx in range(len(idList)):
    genre = returnGenres(idList[idx])
    pop_song_boolean = isPopSong(genre)
    song_index_int = idx + 1
    print(f"Progress => ({(idx+1)}/{(last_idx-first_idx+1)}), {idx+1} : {idList[idx]} , Pop song: {pop_song_boolean}")
    unit_dictionary = {"idx" : song_index_int, "is_pop" : pop_song_boolean, "song_id":idList[idx]}
    resList.append(unit_dictionary)
wb_read.close()

print("Genres Obtained. Now writing to excel sheet... May take awhile...")

wb_write = Workbook()
wb_write = load_workbook(filename=file_name)
ws_write = wb_write["Sheet1"]
print("Workbook opened, writing to sheet...")

for idx in range(len(resList)):
    row_write = resList[idx]["idx"] + 1
    column_write = column_to_write
    genre_write = resList[idx]["is_pop"]
    writeToCell(ws_write,row_write,resList[idx]["is_pop"],column_write)
    print(f"Written Progress => ({idx}/{last_idx}), Row Index: {row_write}, Pop Song: {genre_write}")

print("Saving data onto worksheet... May take awhile...")
wb_write.save(r"C:\Users\ASUS\Desktop\Data collection\tracks.xlsx")

print("Data Saved, Process Complete. Exiting...")
